import re
import asyncio
from datetime import datetime
import logging
import os.path
import aiohttp
import certifi
import requests
from shutil import copyfile
from azure.common.credentials import get_azure_cli_credentials
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter

logger = logging.getLogger(__name__)


def handle_custom_ssl():
    """
    This function is used to inject custom CA certificate into the system.
    It is used to overcome the SSL certificate verification problem
    by injecting locally stored PEM certificate into certifi store
    in virtual env.
    """
    certifi_ca_file = certifi.where()
    certifi_ca_file_original = f"{certifi_ca_file}.orig"
    if os.path.isfile(certifi_ca_file_original):
        # Restore original certifi pem if .orig exists
        copyfile(certifi_ca_file_original, certifi_ca_file)
    else:
        # Backup original certifi pem
        copyfile(certifi_ca_file, certifi_ca_file_original)

    custom_ca_file = "certificate.pem"
    if os.path.isfile(custom_ca_file):
        logger.info(f"Found custom CA file {custom_ca_file} ..")
        custom_ca_file_bytes = open(custom_ca_file, "rb").read()
        venv_ca_path = certifi_ca_file
        logger.info(f"Injecting custom CA into {venv_ca_path} ..")
        with open(venv_ca_path, "ab") as f:
            f.write(custom_ca_file_bytes)
        logger.info("Done ..")


def set_headers(credentials):
    """
    This function is used to set the headers for the requests to the Azure API.
    It takes the credentials object as an argument and returns the headers.
    The headers are used to authenticate the requests to the Azure API.
    The credentials object is used to get the access token.

    Parameters:
        credentials (object): The credentials object is used to get the access token.

    Returns:
        headers (dict):
            The headers are used to authenticate the requests to the Azure API.
    """
    headers = credentials.signed_session().headers
    headers["Content-type"] = "application/json"
    return headers


def get_tenant(headers):
    """
    This function returns the tenant details of the organization.

    Parameters:
        headers (dict): A dictionary of HTTP headers containing the authorization token.

    Returns:
        dict: The tenant detils of the organization.
    """
    try:
        response = requests.get(
            "https://graph.microsoft.com/v1.0/organization", headers=headers, timeout=50
        ).json()
        return response["value"][0]
    except Exception as e:
        raise Exception(e) from e


def get_azure_credentials(endpoint):
    """
    This function is used to get the credentials from the Azure CLI.

    Parameters:
        endpoint: The resource to get the credentials for.

    Returns:
        credentials: The credentials for the resource.
    """
    try:
        credentials, subscription_id = get_azure_cli_credentials(resource=endpoint)
        return credentials
    except Exception as e:
        raise Exception(e) from e


def get_auth_user_details(headers, endpoint, api_version="beta"):
    """
    This function takes a list of user ids and returns a list of dictionaries
    containing the user's MFA registration details.

    Parameters:
        headers: A dictionary containing the authorization header.
        endpoint: The endpoint to query.
        user_ids: A list of user ids.
        api_version: The version of the API to query.

    Returns:
        A list of dictionaries containing the user's MFA registration details.
    """
    try:
        params = {"api-version": api_version}
        response = requests.get(
            f"{endpoint}/myorganization/activities/authenticationMethodUserDetails",
            headers=headers,
            params=params,
            timeout=50,
        )
        if hasattr(response, "json"):
            response = response.json()
        else:
            raise Exception(response.content)
        if "value" not in response:
            raise Exception(response)
        response = response["value"]
        return response
    except Exception as e:
        raise Exception(e) from e


async def get_url(url, user_id, headers, session, params):
    """
    This generic function is used to get the data from the API
    using parallel aiohttp library.

    Parameters:
        url: The url of the API.
        headers: The headers of the API.
        session: The session of the API.
        params: The parameters of the API.

    Returns:
        The data from the API.
    """
    try:
        async with session.get(url=url, headers=headers, params=params) as response:
            response = await response.json()
            # Handle Windows AD accounts.
            if "id" not in response:
                response["id"] = user_id
            if "signInActivity" not in response:
                response["signInActivity"] = {}
                response["signInActivity"]["lastSignInDateTime"] = None
                response["signInActivity"]["lastNonInteractiveSignInDateTime"] = None

            return response

    except Exception as e:
        raise Exception(e) from e


async def get_aad_users(headers, endpoint, user_ids, query_select, api_version="beta"):
    """
    This function takes a list of user ids and returns a list of dictionaries
    containing the user's details.

    Parameters:
        headers: A dictionary containing the authorization header.
        endpoint: The endpoint to query.
        user_ids: A list of user ids.
        api_version: The version of the API to query.

    Returns:
        A list of dictionaries containing the user's details.
    """
    connector = aiohttp.TCPConnector(
        limit=10, force_close=True, enable_cleanup_closed=True
    )
    async with aiohttp.ClientSession(connector=connector) as session:
        user_details = await asyncio.gather(
            *[
                get_url(
                    f"{endpoint}/{api_version}/users/{user_id}",
                    user_id,
                    headers,
                    session,
                    params={"$select": ",".join(query_select)},
                )
                for user_id in user_ids
            ]
        )
        return user_details


def item_to_string(item):
    """
    Converts an item to a string.

    Parameters
    ----------
    item : object
        The item to convert to a string.

    Returns
    -------
    str
        The item converted to a string.

    Notes
    -----
    If the item is None, then "N/A" is returned.
    If the item is True, then "Yes" is returned.
    If the item is False, then "No" is returned.
    If the item is a string, then the string is returned.
    If the item is a datetime,
    then the string representation of the datetime is returned.
    """
    if item is None:
        return "N/A"
    if item is True:
        return "Yes"
    if item is False:
        return "No"
    if isinstance(item, str):
        try:
            return is_datetime(item)
        except ValueError:
            return item
    return item


def is_external(item):
    """
    This function takes a string as an argument and returns a string.
    The string is either "True" or "False".

    The function checks if the string contains the "#EXT#" substring
    which indicates AAD guest user.
    """
    item = bool(re.search("#EXT#", item))
    return item_to_string(item)


def is_datetime(item):
    """
    This function takes a string and returns a datetime object if the string
    is a valid date/time representation. Otherwise it returns "Never".

    Parameters:
        item (str): A string representing a date/time.

    Returns:
        str: A string representing a date/time, or "Never".
    """
    if item and item is not None:
        date = datetime.strptime(item, "%Y-%m-%dT%H:%M:%SZ")
        if date.year > 1999:
            return date.isoformat()
        return "Never"
    return "N/A"


def is_external_domain(item):
    """
    This function takes a string as input and returns the external domain name.
    It uses a regular expression to extract the guest account external domain name.
    It returns "N/A" if the URL is not a valid external guest account domain.

    Parameters:
        item (str): The string to be processed.

    Returns:
        str: The domain name of the URL.
    """
    item = re.findall("(?<=_)(.*)(?=#EXT#)", item)
    if item:
        return item[0]
    return "N/A"


def get_tenant_domain(item):
    """
    This function takes a string as input and returns the tenant domain name.
    It uses a regular expression to extract the tenant domain name user is member of.
    It returns "N/A" if the parameter is not a valid tenant domain.

    Parameters:
        item (str): The string to be processed.

    Returns:
        str: The domain name of the URL.
    """
    item = re.findall("(?<=@)(.*)(?=$)", item)
    if item:
        return item[0]
    return "N/A"


def get_mfa_methods(item):
    """
    This function returns the MFA methods configured for a user.
    It accepts a single argument, which is the user object.
    It returns a string with the MFA methods separated by a comma.
    If no MFA methods are configured, it returns "No AAD MFA configured".
    """
    if item:
        return ",".join(set(item))
    return "No AAD MFA configured"


def xlsx_dict_prep(data):
    """
    This function takes a list of dictionaries and returns a list of dictionaries.
    The output list of dictionaries has the following keys:
    - userId
    - isEnabled
    - userDisplayName
    - userPrincipalName
    - isExternal
    - externalDomain
    - externalUserState
    - externalUserStateLastChangeUTC
    - tenantDomain
    - methodsRegistered
    - onPremisesSyncEnabled
    - lastInteractiveSignInUTC
    - lastNonInteractiveSignInUTC
    The input list of dictionaries is expected to be a list of dictionaries
    representing the merged output of the `get_aad_users` and
    `get_auth_user_details` functions.
    The output list of dictionaries is a list of dictionaries, each representing
    a single user.
    """
    return [
        {
            "userId": x["id"],
            "isEnabled": item_to_string(x["accountEnabled"]),
            "userDisplayName": x["userDisplayName"],
            "userPrincipalName": x["userPrincipalName"],
            "isExternal": is_external(x["userPrincipalName"]),
            "externalDomain": is_external_domain(x["userPrincipalName"]),
            "externalUserState": item_to_string(x["externalUserState"]),
            "externalUserStateLastChangeUTC": item_to_string(
                x["externalUserStateChangeDateTime"]
            ),
            "tenantDomain": get_tenant_domain(x["userPrincipalName"]),
            "methodsRegistered": get_mfa_methods(x["methodsRegistered"]),
            "onPremisesSyncEnabled": item_to_string(x["onPremisesSyncEnabled"]),
            "lastInteractiveSignInUTC": is_datetime(
                x["signInActivity"]["lastSignInDateTime"]
            ),
            "lastNonInteractiveSignInUTC": is_datetime(
                x["signInActivity"]["lastNonInteractiveSignInDateTime"]
            ),
        }
        for x in data
    ]


def adjust_column_width(sheet):
    """
    Adjust the column width of a worksheet.

    :param sheet: The worksheet to be adjusted.
    :type sheet: openpyxl.worksheet.worksheet.Worksheet

    This function takes a sheet as an argument and adjusts the width of each column
    to fit the contents of that column.
    Uses the column_letter function from the openpyxl module to get the
    column letter of each column.
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except ValueError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width


def generate_xlsx(data, sheet_title, table_style, filename):
    """
    Generate an Excel file from a list of dictionaries.

    :param data: A list of dictionaries.
    :param sheet_title: The title of the Excel sheet.
    :param table_style: The name of the Excel table style.
    :param filename: The name of the Excel file.

    :return: None

    Example:

    >>> data = [{'name': 'John', 'age': 20}, {'name': 'Brian', 'age': 25}]
    >>> generate_xlsx(data, 'My Table', 'Table Style Medium 15', 'my_table.xlsx')
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    headers = list(data[0].keys())
    for i in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=i)
        cell.value = headers[i - 1]
    for i in range(2, len(data) + 2):
        for j in range(1, len(headers) + 1):
            cell = sheet.cell(row=i, column=j)
            cell.value = data[i - 2][headers[j - 1]]
    xlsx_header_letter = get_column_letter(len(headers))
    xlsx_header_number = len(data) + 1
    tab = Table(
        displayName=sheet_title.replace(" ", "_").lower(),
        ref=f"A1:{xlsx_header_letter}{xlsx_header_number}",
    )
    style = TableStyleInfo(
        name=table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    adjust_column_width(sheet)
    workbook.save(filename)
